"""
飞书机器人处理器
"""

from typing import Dict, List, Tuple
from loguru import logger
from lark_oapi.api.im.v1 import *
from pathlib import Path
import re
from anthropic import Anthropic


class FeishuBot:
    """飞书机器人"""

    def __init__(self, config, kb, classifier, template_mgr):
        self.config = config
        self.kb = kb
        self.classifier = classifier
        self.template_mgr = template_mgr
        self.app_id = config.bots["feishu"].app_id
        from bots.conversation_manager import ConversationManager
        self.conversation_mgr = ConversationManager(config)

        # 初始化 SDK 客户端用于发送消息和调用 API
        import lark_oapi as lark
        self.lark_client = lark.Client.builder() \
            .app_id(self.app_id) \
            .app_secret(config.bots["feishu"].app_secret) \
            .build()

        # 初始化 LLM client（用于知识库回答生成）
        client_kwargs = {"api_key": config.llm.api_key}
        if config.llm.base_url:
            client_kwargs["base_url"] = config.llm.base_url
        self.llm_client = Anthropic(**client_kwargs)

        # 关键词表配置
        keyword_cfg = config.keyword_table if hasattr(config, 'keyword_table') else {}
        self.keyword_base_token = keyword_cfg.get('base_token', '') if isinstance(keyword_cfg, dict) else getattr(keyword_cfg, 'base_token', '')
        self.keyword_table_id = keyword_cfg.get('table_id', '') if isinstance(keyword_cfg, dict) else getattr(keyword_cfg, 'table_id', '')

        # 加载关键词回复映射
        self.keyword_replies: List[Tuple[str, str]] = []
        self._load_keyword_replies()

        # 会话状态存储（用于快捷模板识别）
        self.conversations: Dict = {}

    def _load_keyword_replies(self):
        """从飞书多维表格加载关键词回复映射"""
        try:
            # 如果没有配置关键词表，尝试从本地 Excel 加载（兼容旧配置）
            if not self.keyword_base_token or not self.keyword_table_id:
                logger.warning("未配置关键词多维表格，尝试从本地 Excel 加载")
                self._load_keyword_replies_from_excel()
                return

            # 使用 httpx 直接调用飞书 API（SDK 版本兼容性问题）
            import httpx
            import os

            app_id = os.environ.get("FEISHU_APP_ID") or self.config.bots["feishu"].app_id
            app_secret = os.environ.get("FEISHU_APP_SECRET") or self.config.bots["feishu"].app_secret

            logger.info(f"开始请求飞书接口获取关键词(table={self.keyword_table_id})，若网络代理有冲突可能在此卡死(最高超时30秒)...")

            # 获取 tenant_access_token
            token_url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
            token_resp = httpx.post(token_url, json={"app_id": app_id, "app_secret": app_secret}, timeout=30, verify=False)
            token_data = token_resp.json()
            token = token_data.get("tenant_access_token")

            if not token:
                logger.error(f"获取飞书 token 失败: {token_data}")
                self._load_keyword_replies_from_excel()
                return

            # 获取关键词表数据
            records_url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{self.keyword_base_token}/tables/{self.keyword_table_id}/records?page_size=500"
            records_resp = httpx.get(records_url, headers={"Authorization": f"Bearer {token}"}, timeout=30, verify=False)
            records_data = records_resp.json()

            if records_data.get("code") != 0:
                logger.error(f"获取飞书关键词表失败: {records_data.get('msg')}")
                self._load_keyword_replies_from_excel()
                return

            items = records_data.get("data", {}).get("items", [])
            loaded_count = 0

            for item in items:
                fields = item.get("fields", {})

                # 字段名：关键词、回复内容、添加状态
                keywords_str = fields.get("关键词", "")
                reply_content = fields.get("回复内容", "")
                status = fields.get("添加状态", "")

                # 只加载状态为「已完成」的条目
                if keywords_str and reply_content and status == "已完成":
                    formatted_reply = self._format_reply_content(str(reply_content))
                    keywords_raw = str(keywords_str).strip()

                    # 判断关键词格式：
                    # 1. 包含顿号「、」 -> 拆分成多个独立关键词
                    # 2. 不包含顿号 -> 作为整体，同时添加常见子串匹配
                    if "、" in keywords_raw:
                        # 顿号分隔格式：拆分成多个关键词
                        keywords_list = [kw.strip() for kw in keywords_raw.split("、") if kw.strip()]
                        for kw in keywords_list:
                            self.keyword_replies.append((kw.lower(), formatted_reply))
                        loaded_count += 1
                    else:
                        # 直接拼接格式：作为整体关键词
                        self.keyword_replies.append((keywords_raw.lower(), formatted_reply))
                        loaded_count += 1

                        # 对于较长的关键词，添加常见子串匹配
                        common_splits = ["帮助文档", "常见问答", "FAQ", "faq", "下载", "官网", "alpha", "Alpha"]
                        for split in common_splits:
                            if split.lower() in keywords_raw.lower() and split.lower() != keywords_raw.lower():
                                self.keyword_replies.append((split.lower(), formatted_reply))

            logger.info(f"从飞书多维表格加载 {loaded_count} 个关键词回复规则（共 {len(self.keyword_replies)} 条匹配规则）")

        except Exception as e:
            logger.error(f"加载飞书关键词失败: {e}")
            # 降级到本地 Excel
            self._load_keyword_replies_from_excel()

    def _load_keyword_replies_from_excel(self):
        """从本地 Excel 加载关键词回复映射（降级方案）"""
        try:
            from openpyxl import load_workbook

            excel_path = Path(__file__).parent.parent.parent / "data" / "documents" / "list.xlsx"
            if not excel_path.exists():
                logger.warning(f"关键词回复文件不存在: {excel_path}")
                return

            wb = load_workbook(excel_path)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                keywords_str = row[1]  # 关键词列
                reply_content = row[2]  # 回复内容列
                status = row[4]  # 添加状态列

                if keywords_str and reply_content and status == "已完成":
                    keywords = [kw.strip() for kw in str(keywords_str).split("、") if kw.strip()]
                    formatted_reply = self._format_reply_content(str(reply_content))
                    for kw in keywords:
                        self.keyword_replies.append((kw.lower(), formatted_reply))

            logger.info(f"从本地 Excel 加载 {len(self.keyword_replies)} 个关键词回复规则")

        except Exception as e:
            logger.error(f"加载本地关键词回复映射失败: {e}")

    def _format_reply_content(self, content: str) -> str:
        """
        深度格式化回复内容，优化排版和换行
        """
        import re

        # A. 基础清理
        content = re.sub(r'[ \t]+', ' ', content)  # 合并空格
        
        # B. 修复单词被意外折断（例如 OpenAPI/ \n Swagger -> OpenAPI/Swagger）
        content = re.sub(r'([a-zA-Z0-9/\-_]+)\s*\n\s*([a-zA-Z0-9/\-_]+)', r'\1\2', content)

        # C. 编号标题处理 (例如 "1.使用场景 问题描述" -> "1.使用场景\n问题描述")
        # 匹配 1.标题 后面跟着非换行内容的情况
        content = re.sub(r'(\d+\.[\u4e00-\u9fa5]{2,10})\s+([^\n])', r'\n\1\n\2', content)
        # 即使没有空格也尝试切分关键标题
        titles = ["使用场景", "常见问题", "注意事项", "快速开始", "相关链接"]
        for title in titles:
            content = re.sub(f'(\\d+\\.{title})', r'\n\1\n', content)

        # D. 特定句子开头前换行
        content = re.sub(r'([。？?！!])\s*([1-9]\.)', r'\1\n\2', content)
        content = re.sub(r'\s+(无论|如果|提示|注意)', r'\n\1', content)

        # E. 链接项处理：支持长描述 + 自动识别“立即参与内测”等引导语
        # 第一步：处理标准的 标题 ：链接
        link_pattern = r'([\u4e00-\u9fa5\(\)（）\w\s\-\.]{2,100}?)\s*(?:☞|：|:)\s*(https?://[^\s\n\u4e00-\u9fa5]+)'
        
        def replace_link(match):
            name = match.group(1).strip()
            name = name.replace('\n', '').replace('\r', '')
            url = match.group(2).strip()
            # 使用 ☞ 符号显得更美观且专业
            return f'\n{name} ☞ {url}\n'
            
        content = re.sub(link_pattern, replace_link, content)

        # F. 第二步：处理孤立的链接前引导语（如“立即参与内测：http...”）
        content = re.sub(r'(立即参与\w*|点击\w*|查看\w*)\s*(?:：|:)\s*(https?://[^\s\n]+)', r'\n\1 ☞ \2\n', content)

        # G. 特殊修复：如果 URL 后面直接跟着下一个标题，强制双换行
        content = re.sub(r'(https?://[^\s\n]+)\s+([\u4e00-\u9fa5\d]{2,})', r'\1\n\n\2', content)

        # H. 收尾工作
        content = re.sub(r'\n{3,}', '\n\n', content)
        content = content.strip()
        
        # 确保每行开头的 1. 2. 3. 干净
        content = re.sub(r'\n\s*(\d+\.)', r'\n\1', content)

        return content

    async def _try_answer_from_kb(self, chat_id: str, question: str) -> Tuple[bool, str]:
        """
        尝试从知识库回答问题

        Args:
            chat_id: 所在群聊/对话的ID
            question: 用户问题

        Returns:
            (can_answer, answer_text)
        """
        has_history = hasattr(self, 'chat_history') and chat_id in self.chat_history and len(self.chat_history[chat_id]) > 0
        try:
            # 检索知识库
            docs = await self.kb.search(question)

            if not docs and not has_history:
                logger.info(f"知识库无匹配结果且无历史语境: {question}")
                return (False, "", [])

            # 有文档或有历史聊天，交给 LLM 回答
            context = ""
            if docs:
                logger.info(f"知识库找到 {len(docs)} 个相关文档")
                context = await self.kb.search_with_context(question)
                
            answer = await self._generate_answer_from_context(chat_id, question, context)

            return (True, answer, docs)

        except Exception as e:
            logger.error(f"知识库检索失败: {e}")
            return (False, "", [])

    async def _generate_answer_from_context(self, chat_id: str, question: str, context: str) -> str:
        """
        基于知识库内容和历史上下文生成回答

        Args:
            chat_id: 所在群聊/对话的ID
            question: 用户问题
            context: 知识库检索到的上下文
        """
        if not hasattr(self, 'chat_history'):
            self.chat_history = {}
        if chat_id not in self.chat_history:
            self.chat_history[chat_id] = []
            
        # 强化 Apifox 品牌意识，严禁提及 Apidog
        prompt = f"""你是 Apifox 技术支持助手，解答用户关于 Apifox 产品的技术及使用问题。
        
【强制约束】
1. 严禁在任何回复中提及 "Apidog" 关键词。
2. 如果提供的上下文片段中包含 "Apidog"，请在回答时将其自动替换为 "Apifox" 或根据语境合理翻译，绝对不能输出 Apidog 原词。
3. 只提供 Apifox 相关的技术支持。
4. 直接回答问题，严禁在回复开头使用 "核心解答："、"回答："、"结论：" 等任何前缀。
5. 回复应当自然、专业，像技术支持人员在对话。"""
        if context:
            prompt += f"\n\n请参考以下检索到的文档片段：\n{context}\n\n- 如果文档内容能回答问题，直接给出答案\n- 如果无法完全回答，提供已知信息或链接\n- 不要编造文档外的不实能力"
        prompt += f"\n\n当前用户的新问题：{question}"

        # 加载历史上下文作为 prompt list
        messages = list(self.chat_history[chat_id])
        messages.append({"role": "user", "content": prompt})

        try:
            response = self.llm_client.messages.create(
                model=self.config.llm.model,
                max_tokens=2000,
                temperature=0.3,
                messages=messages
            )

            # 初始化 thinking 缓存
            self._last_thinking_parts = []

            # 提取文本内容（处理 ThinkingBlock，thinking 部分存入 self._last_thinking_parts）
            ai_text = self._extract_text_from_response(response.content, chat_id)

            # 将思考过程转发到 bot 测试群（无论来自哪个群）
            if self._last_thinking_parts:
                import asyncio
                asyncio.ensure_future(
                    self._relay_thinking_to_test_group(chat_id, question, self._last_thinking_parts)
                )
            
            # 更新历史记忆（仅保存问题原意，不存大段 Prompt 节约 token）
            self.chat_history[chat_id].append({"role": "user", "content": question})
            self.chat_history[chat_id].append({"role": "assistant", "content": ai_text})
            
            # 保留最近的5轮对话（10条消息）
            if len(self.chat_history[chat_id]) > 10:
                self.chat_history[chat_id] = self.chat_history[chat_id][-10:]
                
            return ai_text

        except Exception as e:
            logger.error(f"LLM 生成回答失败: {e}")
            return "抱歉，生成回答时出现错误，请稍后再试。"

    # Bot 测试群 chat_id（显示思考内容）
    BOT_TEST_CHAT_ID = "oc_9181c1f2869b0ec0af937f5c6f9a82aa"

    def _extract_text_from_response(self, content_blocks, chat_id: str = "") -> str:
        """
        从 LLM 响应中提取文本内容
        处理 TextBlock 和 ThinkingBlock 混合的情况
        仅返回 text 块内容，thinking 块通过 _relay_thinking_to_test_group 异步转发
        """
        text_parts = []
        thinking_parts = []

        for block in content_blocks:
            block_type = getattr(block, 'type', None)
            if block_type == 'text':
                text_parts.append(block.text)
            elif block_type == 'thinking':
                thinking_text = getattr(block, 'thinking', '') or getattr(block, 'text', '')
                if thinking_text:
                    thinking_parts.append(thinking_text)
            elif hasattr(block, 'text'):
                text_parts.append(block.text)
            else:
                text_parts.append(str(block))

        # 把 thinking 存到实例变量，供 _generate_answer_from_context 转发
        self._last_thinking_parts = thinking_parts
        return "\n".join(text_parts)

    async def _relay_thinking_to_test_group(self, source_chat_id: str, question: str, thinking_parts: list):
        """
        将 LLM 思考过程转发到 bot 测试群。
        无论问题从哪个群发出，思考内容都会在测试群可见。
        如果本身就是测试群，此方法也仍然执行（避免遗漏）。
        """
        if not thinking_parts:
            return
        try:
            thinking_text = "\n\n".join(thinking_parts)
            # 截断过长的思考内容以免消息过大
            if len(thinking_text) > 3000:
                thinking_text = thinking_text[:3000] + "\n…（思考内容过长已截断）"

            is_same_group = source_chat_id == self.BOT_TEST_CHAT_ID
            group_label = "（同一群）" if is_same_group else f"（来自群 `{source_chat_id}`）"
            header = f"💭 **LLM 思考过程** {group_label}\n❓ 问题：{question[:60]}{'…' if len(question) > 60 else ''}"
            relay_content = f"{header}\n\n{thinking_text}"

            await self.send_message(self.BOT_TEST_CHAT_ID, relay_content)
            logger.info(f"已将思考过程转发到 bot 测试群，来源群: {source_chat_id}")
        except Exception as e:
            logger.warning(f"转发思考过程到测试群失败（非致命）: {e}")

    async def handle_message(self, event: Dict):
        """
        处理飞书消息
        """
        try:
            chat_id = event.get("chat_id", "")
            user_id = event.get("sender", {}).get("sender_id", {}).get("user_id", "")
            content = event.get("content", {})
            message_id = event.get("message_id", "")

            # 解析内容
            if isinstance(content, str):
                import json
                content = json.loads(content)
            text_content = content.get("text", "")

            # 记录历史
            conversation_key = chat_id
            is_active = self.conversation_mgr.is_conversation_active(conversation_key)
            
            if is_active:
                self.conversation_mgr.add_to_history(conversation_key, "user", text_content)
                await self._handle_conversation_step(chat_id, user_id, text_content)
                return

            # 如果没有活跃对话，匹配关键词
            matched_reply = self._match_keyword_reply(text_content)
            if matched_reply:
                await self.send_message(chat_id, matched_reply)
                return

            # 触发条件：必须被 @
            if self._is_bot_mentioned(event):
                await self._handle_new_question(chat_id, user_id, text_content, message_id)

        except Exception as e:
            logger.error(f"处理飞书消息失败: {e}")
    def _is_bot_mentioned(self, event: Dict) -> bool:
        """检查机器人是否被 @"""
        mentions = event.get("message", {}).get("mentions", [])
        known_bot_open_id = "ou_64e28296caaaa7841cec5af1ecbd8f9a"
        for mention in mentions:
            mention_id = mention.get("id", "")
            if mention_id == self.app_id or mention_id == known_bot_open_id:
                return True
        return False

    def _match_keyword_reply(self, text: str) -> str:
        """匹配关键词并返回回复"""
        text_lower = text.lower()
        for keyword, reply in self.keyword_replies:
            if keyword in text_lower:
                return reply
        return ""

    def _contains_trigger_keywords(self, text: str) -> bool:
        """检查是否包含触发关键字"""
        feishu_cfg = self.config.bots.get("feishu", {})
        if isinstance(feishu_cfg, dict):
            triggers = feishu_cfg.get("triggers", {})
            keywords = triggers.get("keywords", [])
        else:
            triggers = getattr(feishu_cfg, "triggers", {})
            keywords = triggers.get("keywords", [])
        return any(keyword in text for keyword in keywords) if keywords else False

    async def _handle_new_question(self, chat_id: str, user_id: str, question: str, message_id: str = ""):
        """处理新问题：三段式反馈流程"""
        try:
            # 阶段1：收到 (GET)
            if message_id:
                logger.info(f"收到新问题 {message_id}，回复 OK...")
                await self._send_reaction(message_id, "OK")

            import re
            # 清理飞书 mention 标签，只删除 @_xxx 格式的 mention，保留消息内容
            # 飞书消息格式可能包含：@_user_xxx 或 @_all 等 mention
            clean_question = re.sub(r'@_[a-zA-Z0-9_]+\s*', '', question).strip()
            logger.debug(f"原始问题: {question[:100]}, 清理后: {clean_question[:100] if clean_question else '空'}")

            # 如果清理后为空，使用原始问题（去掉可能的 at 标签）
            if not clean_question:
                # 飞书富文本格式：<at user_id="xxx">名字</at>
                clean_question = re.sub(r'<at[^>]*>.*?</at>', '', question).strip()
                # 再尝试去掉纯 @ 提示
                clean_question = re.sub(r'@\S+\s*', '', clean_question).strip()
                # 如果还是空，直接用原始问题去掉空白
                if not clean_question and question.strip():
                    clean_question = question.strip()
                    logger.warning(f"问题清理失败，使用原始内容: {clean_question[:50]}")

            # 优先从知识库回答
            can_answer, answer, docs = await self._try_answer_from_kb(chat_id, clean_question)
            if can_answer:
                from src.utils.message_card_builder import build_faq_answer_card
                # 将 langchain 文档对象转换为展示用的列表
                related_docs = []
                if docs:
                    for d in docs[:3]: # 最多取3条
                        title = d.metadata.get('title') or d.metadata.get('source', '文档')
                        # 如果 title 太长则截断
                        if len(title) > 20: title = title[:20] + "..."
                        url = d.metadata.get('url', 'https://help.apifox.com')
                        related_docs.append({"title": title, "url": url})
                
                card_json = build_faq_answer_card(clean_question, answer, related_docs=related_docs)
                await self._send_card(chat_id, card_json)
                # 阶段3：处理完毕
                if message_id:
                    await self._send_reaction(message_id, "DONE")
                return

            # 知识库没找到，进入深度分析阶段，发送“在做了”
            if message_id:
                await self._send_reaction(message_id, "OK")

            # 知识库没找到，先分类问题类型
            context = await self.kb.search_with_context(clean_question) if self.kb.vectorstore else "知识库未初始化"
            classification = await self.classifier.classify(clean_question, context, {})
            question_type = classification.get("type", "unknown")
            suggested_answer = classification.get("suggested_answer", "")

            # 兜底逻辑：增强型分类判定
            low_confidence = classification.get("confidence", 0) < 0.6
            
            # 1. 需求判断关键词
            feature_keywords = ["建议", "能不能", "需求", "希望", "导出", "想要", "期待"]
            # 2. Bug 判断关键词
            bug_keywords = ["报错", "无法", "500", "404", "错误", "失败", "异常", "崩溃"]

            if question_type == "unknown" or low_confidence:
                if any(k in clean_question.lower() for k in bug_keywords):
                    question_type = "bug"
                    logger.info(f"由于检测到 Bug 强相关核心词，强制分类为 bug")
                elif any(k in clean_question.lower() for k in feature_keywords):
                    question_type = "feature"
                    logger.info(f"分类不确定，但检测到需求关键词，修正为 feature")

            logger.info(f"问题分类: type={question_type}, confidence={classification.get('confidence')}")

            # 根据问题类型采取不同策略
            if question_type == "usage":
                # 使用问题：知识库没找到时，直接告知用户并记录为 FAQ
                # 不启动缺陷收集流程
                if suggested_answer:
                    await self.send_message(chat_id, suggested_answer)
                else:
                    await self.send_message(chat_id,
                        f"抱歉，我在帮助文档中暂时没有找到关于「{clean_question[:30]}」的内容。\n\n"
                        f"建议您查阅官方文档：\n"
                        f"- Apifox: https://help.apifox.com\n\n"
                        f"如有其他问题，请继续提问。"
                    )
                # 处理完毕
                if message_id:
                    await self._send_reaction(message_id, "DONE")
                return

            # Bug 或 Feature：启动智能信息收集流程
            conversation_key = chat_id

            # 根据类型选择收集字段
            if question_type == "feature":
                target_fields = ["scenario", "background"]
            else:
                # Bug 类型
                target_fields = ["version", "os", "steps"]

            # 初始实体提取
            history_for_extraction = [{"role": "user", "content": clean_question}]
            extracted = await self.classifier.extract_info(history_for_extraction, target_fields)

            self.conversation_mgr.start_conversation(conversation_key, clean_question)
            self.conversation_mgr.add_to_history(conversation_key, "user", clean_question)
            self.conversation_mgr.update_extracted_data(conversation_key, extracted)

            # 记录问题类型，后续创建工单时使用
            self.conversation_mgr.conversations[conversation_key]["forced_type"] = question_type

            conversation = self.conversation_mgr.conversations[conversation_key]
            collected = conversation["collected_data"]
            missing = [f for f in target_fields if not collected.get(f) or collected.get(f) == "未提及"]

            from src.utils.message_card_builder import build_smart_collection_card

            # 发送信息收集卡片
            card_json = build_smart_collection_card(
                clean_question,
                collected,
                missing,
                state="collecting",
                card_type=question_type
            )

            new_msg_id = await self._send_card(chat_id, card_json)
            if new_msg_id:
                self.conversation_mgr.set_progress_message_id(conversation_key, new_msg_id)
            # 阶段3：处理完毕
            if message_id:
                await self._send_reaction(message_id, "DONE")

        except Exception as e:
            logger.error(f"处理新问题失败: {e}")

    async def _handle_conversation_step(self, chat_id: str, user_id: str, text: str):
        """处理信息收集步进"""
        conversation_key = chat_id
        conversation = self.conversation_mgr.conversations[conversation_key]
        
        # 使用 LLM 提取当前进度
        target_fields = ["version", "os", "steps"] if conversation.get("forced_type") != "feature" else ["scenario", "background"]
        extracted = await self.classifier.extract_info(conversation["history"], target_fields)
        
        # 更新已收集数据
        self.conversation_mgr.update_extracted_data(conversation_key, extracted)
        collected = conversation["collected_data"]
        
        # 检查是否收集完成
        missing = [f for f in target_fields if not collected.get(f) or collected.get(f) == "未提及"]
        question_type = conversation.get("forced_type", "bug")
        
        from src.utils.message_card_builder import build_smart_collection_card
        
        if not missing or len(conversation["history"]) >= 6:
            # 进入确认阶段
            self.conversation_mgr.set_confirming(conversation_key)
            
            # 如果是需求类型，使用专门的正式需求确认卡片
            if conversation.get("forced_type") == "feature":
                from src.utils.message_card_builder import build_formal_feature_card
                card_json = build_formal_feature_card(
                    conversation["initial_question"],
                    collected.get("scenario", ""),
                    collected.get("background", ""),
                    state="confirming"
                )
            else:
                card_json = build_smart_collection_card(
                    conversation["initial_question"],
                    collected,
                    missing,
                    state="confirming",
                    card_type=question_type
                )
        else:
            # 继续收集，更新进度卡片
            card_json = build_smart_collection_card(
                conversation["initial_question"],
                collected,
                missing,
                state="collecting",
                card_type=question_type
            )

        # 更新或发送进度卡片
        existing_msg_id = self.conversation_mgr.get_progress_message_id(conversation_key)
        if existing_msg_id:
            await self._update_message_card(existing_msg_id, card_json)
        else:
            new_msg_id = await self._send_card(chat_id, card_json)
            if new_msg_id:
                self.conversation_mgr.set_progress_message_id(conversation_key, new_msg_id)

    async def handle_card_action(self, event: Dict) -> Dict:
        """处理卡片交互"""
        action = event.get("action", {})
        value = action.get("value", {})
        user_id = event.get("operator", {}).get("user_id", "")
        message_id = event.get("message_id", "")
        
        logger.info(f"收到卡片动作: {value}")
        
        action_type = value.get("action")
        
        if action_type == "confirm_ticket":
            return await self._handle_confirm_ticket(event)
        elif action_type == "edit_ticket":
            return {"toast": {"type": "info", "content": "好的，请直接在群内发送内容，我将自动补充或修正信息。"}}
        elif action_type == "input_field":
            return {"toast": {"type": "info", "content": "请直接在群内回复文字或截图即可，我会自动提取并更新。"}}
        elif action_type == "incomplete_submit":
            return {"toast": {"type": "warning", "content": "信息尚未收集完整，请先根据卡片提示在群内补充信息。"}}
        elif action_type == "force_submit":
            # 强制提交：即使信息不完整也创建工单
            return await self._handle_confirm_ticket(event)

        return {}

    async def _handle_confirm_ticket(self, event: Dict) -> Dict:
        """完成并创建工单"""
        message_id = event.get("message_id", "")
        user_id = event.get("operator", {}).get("user_id", "")
        
        # 这里需要找到对应的会话
        # 在真实场景中，我们可以通过 card value 携带 session_id
        # 此处简化处理：假设 message_id 映射到某个活跃会话
        session_found = None
        for key, conv in self.conversation_mgr.conversations.items():
            if conv.get("progress_message_id") == message_id:
                session_found = (key, conv)
                break
        
        if not session_found:
            return {"toast": {"type": "error", "content": "会话已过期"}}
            
        key, conv = session_found
        collected = conv["collected_data"]
        
        # 确定类型并创建工单
        forced_type = conv.get("forced_type")
        if not forced_type:
            context = await self.kb.search_with_context(conv["initial_question"])
            classification = await self.classifier.classify(conv["initial_question"], context, collected)
            ticket_type = classification["type"]
        else:
            ticket_type = forced_type
            
        # 创建工单（复用 lark-cli wrapper）
        from src.integrations.lark_cli_wrapper import LarkCliWrapper
        lark_cli = LarkCliWrapper(self.config)
        
        if ticket_type == "feature":
            res = lark_cli.create_feature_record(collected, user_id)
        else:
            # 兼容字段名
            bug_data = {
                "title": collected.get("title", "未命名Bug"),
                "version": collected.get("version", "未知"),
                "environment": collected.get("os", "未知"),
                "steps": collected.get("steps", "未提及")
            }
            res = lark_cli.create_bug_record(bug_data, user_id)
            
        if res.get("success"):
            # 更新卡片为已完成状态
            from src.utils.message_card_builder import MessageCardBuilder
            builder = MessageCardBuilder()
            builder.set_header("工单创建成功", "success")
            builder.add_div(f"✅ **已提交工单**\n查看地址：{res.get('url')}")
            await self._update_message_card(message_id, builder.to_json())
            
            # 清理会话
            self.conversation_mgr.cancel_conversation(key)
            return {"toast": {"type": "success", "content": "提交成功"}}
        else:
            return {"toast": {"type": "error", "content": f"提交失败: {res.get('error')}"}}

    async def _send_card(self, chat_id: str, card_json: str) -> Optional[str]:
        """发送交互式卡片"""
        import json
        request = CreateMessageRequest.builder() \
            .receive_id_type("chat_id") \
            .request_body(CreateMessageRequestBody.builder() \
                .receive_id(chat_id) \
                .msg_type("interactive") \
                .content(card_json) \
                .build()) \
            .build()
        response = self.lark_client.im.v1.message.create(request)
        if response.success():
            logger.info(f"✅ 成功发送交互式卡片到 {chat_id}")
            return response.data.message_id
        logger.error(f"发送卡片失败: {response.msg}")
        return None

    async def _send_reaction(self, message_id: str, emoji_type: str = "OK"):
        """对消息发送表情回复，表示正在处理"""
        try:
            import httpx
            import ssl
            # 创建不验证 SSL 的上下文，绕过企业代理证书问题
            ssl_ctx = ssl.create_default_context()
            ssl_ctx.check_hostname = False
            ssl_ctx.verify_mode = ssl.CERT_NONE
            async with httpx.AsyncClient(verify=False, timeout=10.0) as client:
                # 获取 token
                token_url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
                token_resp = await client.post(
                    token_url,
                    json={"app_id": self.app_id, "app_secret": self.config.bots["feishu"].app_secret}
                )
                token = token_resp.json().get("tenant_access_token")
                if not token:
                    logger.warning(f"获取 token 失败，无法发送 reaction: {emoji_type}")
                    return

                # 发送表情
                reaction_url = f"https://open.feishu.cn/open-apis/im/v1/messages/{message_id}/reactions"
                resp = await client.post(
                    reaction_url,
                    headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
                    json={"reaction_type": {"emoji_type": emoji_type}}
                )
                result = resp.json()
                if result.get("code") == 0:
                    logger.info(f"✅ 已发送表情回复: {emoji_type}")
                else:
                    logger.warning(f"发送表情回复响应异常: {result}")
        except Exception as e:
            logger.warning(f"发送表情回复失败（非致命）: {e}")

    async def _update_message_card(self, message_id: str, card_json: str):
        """更新交互式卡片 (PATCH)"""
        import httpx
        try:
            async with httpx.AsyncClient() as client:
                # 获取 token
                token_url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
                token_resp = await client.post(token_url, json={"app_id": self.app_id, "app_secret": self.config.bots["feishu"].app_secret})
                token = token_resp.json().get("tenant_access_token")

                # PATCH 消息
                patch_url = f"https://open.feishu.cn/open-apis/im/v1/messages/{message_id}"
                await client.patch(
                    patch_url,
                    headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
                    json={"content": card_json}
                )
        except Exception as e:
            logger.error(f"更新卡片失败: {e}")

    def _translate_type(self, ticket_type: str) -> str:
        """翻译问题类型为中文"""
        type_map = {
            "bug": "缺陷/Bug",
            "feature": "功能需求",
            "usage": "使用问题",
            "unknown": "未知类型"
        }
        return type_map.get(ticket_type, "其他")

    async def _handle_command(
        self,
        chat_id: str,
        user_id: str,
        command: str
    ):
        """处理命令"""
        if command == "/cancel":
            conversation_key = chat_id
            self.conversation_mgr.cancel_conversation(conversation_key)
            if hasattr(self, 'chat_history') and chat_id in self.chat_history:
                self.chat_history[chat_id] = []
            await self.send_message(chat_id, "已取消当前流程并清除了对话上下文记忆。")

        elif command == "/help":
            help_text = """
使用帮助：
1. 直接发送问题，我会引导您收集信息
2. /cancel - 取消当前对话
3. /help - 显示帮助
            """
            await self.send_message(chat_id, help_text)

    async def send_message(self, chat_id: str, content: str):
        """
        发送飞书消息（自动检测链接并转为富文本格式）
        """
        try:
            import json
            import re
            from lark_oapi.api.im.v1 import CreateMessageRequest, CreateMessageRequestBody

            # 检测是否包含链接
            url_pattern = r'https?://[^\s<>\[\]]+'
            urls = re.findall(url_pattern, content)

            if urls:
                rich_content = self._convert_to_rich_text(content)
                content_raw = json.dumps(rich_content, ensure_ascii=False)
                msg_type = "post"
            else:
                content_raw = json.dumps({"text": content}, ensure_ascii=False)
                msg_type = "text"

            request = CreateMessageRequest.builder() \
                .receive_id_type("chat_id") \
                .request_body(CreateMessageRequestBody.builder() \
                    .receive_id(chat_id) \
                    .msg_type(msg_type) \
                    .content(content_raw) \
                    .build()) \
                .build()

            response = self.lark_client.im.v1.message.create(request)

            if not response.success():
                logger.error(f"发送飞书消息失败: {response.code} - {response.msg}")
            else:
                logger.info(f"成功发送飞书消息到 {chat_id}")

        except Exception as e:
            logger.error(f"发送飞书消息异常: {e}")

    def _convert_to_rich_text(self, content: str) -> dict:
        """
        将文本转换为飞书富文本格式
        """
        import re
        url_pattern = r'(https?://[^\s<>\[\]]+)'
        lines = content.split('\n')
        content_blocks = []
        for line in lines:
            line = line.rstrip()
            if not line:
                content_blocks.append([{"tag": "text", "text": ""}])
                continue
            parts = re.split(url_pattern, line)
            line_elements = []
            for part in parts:
                if not part:
                    continue
                if re.match(url_pattern, part):
                    line_elements.append({"tag": "a", "text": part, "href": part})
                else:
                    line_elements.append({"tag": "text", "text": part})
            if line_elements:
                content_blocks.append(line_elements)
        return {"zh_cn": {"content": content_blocks}}

    async def handle_bot_added_to_chat(self, chat_id: str, operator_id: str = ""):
        """处理机器人被添加到群聊"""
        logger.info(f"机器人被添加到群 chat_id: {chat_id}")
        welcome = getattr(self.config.bots.get("feishu", {}), "welcome_message", "")
        if welcome:
            await self.send_message(chat_id, welcome)

    async def handle_bot_removed_from_chat(self, chat_id: str, operator_id: str = ""):
        """处理机器人被移出群聊"""
        logger.info(f"机器人被移出群 chat_id: {chat_id}")
